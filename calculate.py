import pandas as pd
from tabulate import tabulate
from openpyxl.styles import Alignment, Font, PatternFill

class MSISavingsCalculator:
    def __init__(self):
        self.total_amount = 0.0
        self.annual_interest_rate = 0.0
        self.months = 0
        self.monthly_payment = 0.0
        self.output_choice = 'terminal'

    def get_user_input(self):
        """Collect and validate user input."""
        self.total_amount = self.prompt_for_positive_float("Enter the total amount of money $ to finance: ")
        self.annual_interest_rate = self.prompt_for_positive_float("Enter the annual interest rate (e.g., 10 for 10%): ") / 100
        self.months = self.prompt_for_positive_int("Enter the number of months to finance (e.g., 12): ")
        self.monthly_payment = self.total_amount / self.months
        self.output_choice = self.prompt_for_output_choice()

    def prompt_for_positive_float(self, prompt):
        """Prompt the user for a positive floating-point number."""
        while True:
            try:
                value = float(input(prompt))
                if value <= 0:
                    print("Value must be a positive number. Please try again.")
                    continue
                return value
            except ValueError:
                print("Invalid input. Please enter a numeric value.")

    def prompt_for_positive_int(self, prompt):
        """Prompt the user for a positive integer."""
        while True:
            try:
                value = int(input(prompt))
                if value <= 0:
                    print("Value must be a positive integer. Please try again.")
                    continue
                return value
            except ValueError:
                print("Invalid input. Please enter an integer value.")

    def prompt_for_output_choice(self):
        """Prompt the user for their output preference."""
        while True:
            choice = input("Would you like to save the results to an Excel file? (yes/no): ").strip().lower()
            if choice in ('yes', 'y'):
                return 'excel'
            elif choice in ('no', 'n'):
                return 'terminal'
            else:
                print("Invalid choice. Please enter 'yes' or 'no'.")


    def calculate_savings(self):
        """Calculate the savings over the financing period, including total interest without withdrawals."""
        import calendar

        data = []
        daily_interest_rate = self.annual_interest_rate / 365
        current_balance = self.total_amount
        total_interest_earned = 0.0

        # Variables for interest without withdrawals
        current_balance_no_withdrawals = self.total_amount
        total_interest_earned_no_withdrawals = 0.0

        for month in range(1, self.months + 1):
            # Get the actual number of days in the current month
            days_in_month = calendar.monthrange(2021, month % 12 or 12)[1]  # Using 2021 as a non-leap year

            # Calculate interest for the month
            interest_earned = current_balance * daily_interest_rate * days_in_month
            total_interest_earned += interest_earned

            # Calculate interest without withdrawals (compounded)
            interest_earned_no_withdrawals = current_balance_no_withdrawals * daily_interest_rate * days_in_month
            total_interest_earned_no_withdrawals += interest_earned_no_withdrawals
            current_balance_no_withdrawals += interest_earned_no_withdrawals

            # Update balances
            starting_balance = current_balance
            ending_balance = current_balance + interest_earned - self.monthly_payment

            # Append data for the current month
            data.append({
                "Month": month,
                "Starting Balance": round(starting_balance, 2),
                "Interest Earned": round(interest_earned, 2),
                "Ending Balance": round(ending_balance, 2)
            })

            # Prepare for next iteration
            current_balance = ending_balance

        # Simple interest calculation for comparison
        simple_interest = self.total_amount * self.annual_interest_rate

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create a summary dictionary with additional rows
        summary = {
            "Total Amount": f"${round(self.total_amount, 2):,.2f}",
            "Monthly Payment": f"${round(self.monthly_payment, 2):,.2f}",
            "Total Interest Earned": f"${round(total_interest_earned, 2):,.2f}",
            "Interest If Untouched (Compounded)": f"${round(total_interest_earned_no_withdrawals, 2):,.2f}",
            "Interest If Untouched (Simple)": f"${round(simple_interest, 2):,.2f}"
        }

        return df, summary


    def save_to_excel(self, df, summary):
        """Save the DataFrame and summary to an Excel file with the 'Summary' header spanning two columns."""
        output_file = "MSI_Savings_Calculation.xlsx"
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
                # Write the monthly breakdown starting from row 7 (adjusted for extra summary rows)
                df.to_excel(writer, sheet_name='Results', index=False, startrow=len(summary) + 3)
                
                # Access the workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets['Results']
                
                # Merge cells A1:B1 and set "Summary" as the header
                worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
                worksheet['A1'] = 'Summary'
                worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
                worksheet['A1'].font = Font(size=14, bold=True)
                worksheet['A1'].fill = PatternFill("solid", fgColor="cbc1e3")
                
                # Write the summary data underneath the "Summary" header
                for i, (desc, value) in enumerate(summary.items(), start=2):
                    worksheet.cell(row=i, column=1).value = desc
                    worksheet.cell(row=i, column=2).value = value
                    worksheet.cell(row=i, column=1).alignment = Alignment(horizontal='right')
                    worksheet.cell(row=i, column=2).alignment = Alignment(horizontal='center')
                    worksheet.cell(row=i, column=1).font = Font(size=12)
                    worksheet.cell(row=i, column=2).font = Font(size=12)
                
                # Formatting headers of the monthly breakdown
                header_row = len(summary) + 4  # Adjusted for summary rows and spacing
                for cell in worksheet[header_row]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill("solid", fgColor="DDDDDD")
                
                # Center alignment for data cells
                for row in worksheet.iter_rows(min_row=header_row+1, min_col=1, max_col=df.shape[1], max_row=worksheet.max_row):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Adjust column widths
                from openpyxl.utils import get_column_letter
                for col in worksheet.columns:
                    max_length = 0
                    column_index = col[0].column  # Column index as integer
                    column_letter = get_column_letter(column_index)
                    for cell in col:
                        if cell.value:
                            try:
                                length = len(str(cell.value))
                                if length > max_length:
                                    max_length = length
                            except:
                                pass
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            print(f"Data successfully saved to {output_file}")
        except Exception as e:
            print(f"An error occurred while saving to Excel: {e}")



    def display_in_terminal(self, df, summary):
        """Display the summary and DataFrame in the terminal."""
        # Display the summary
        print("\nSummary:")
        summary_df = pd.DataFrame(summary.items(), columns=['Description', 'Value'])
        print(tabulate(summary_df, headers='keys', tablefmt='psql', showindex=False))

        # Display the DataFrame as a table
        print("\nMonthly Breakdown:")
        print(tabulate(df, headers='keys', tablefmt='psql', showindex=False))



    def run(self):
        """Run the calculator."""
        self.get_user_input()
        df, summary = self.calculate_savings()
        if self.output_choice == 'excel':
            self.save_to_excel(df, summary)
        else:
            self.display_in_terminal(df, summary)

if __name__ == "__main__":
    calculator = MSISavingsCalculator()
    calculator.run()
